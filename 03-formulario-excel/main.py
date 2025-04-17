import openpyxl
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import os
nombre_archivo = 'datos.xlsx'


# Comprobar si el archivo existe
if os.path.exists(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
else:
# Crear el libro de excel
    wb = Workbook() # creo el libro
    ws = wb.active # Hoja activa
    ws.append(["Nombre", "Edad", "Email", "Telefono", "Direccion"]) # Creo una nueva fila





def guardar_datos():

    nombre = entry_nombre.get()
    edad = entry_edad.get()
    email = entry_email.get()
    telefono = entry_telefono.get()
    direccion = entry_direccion.get()

    if not nombre or not edad or not email or not telefono or not direccion:
        messagebox.showwarning(title="Advertencia", message="Todos los campos son obligatorios")
        return

    try:
        edad = int(edad)
        telefono = int(telefono)
    except ValueError as err:
        messagebox.showwarning(title="Advertencia", message="Edad y Telefono deben ser numeros")
        return

    # if not re.match(pattern="[^@]+@[^@]+\.[^@]+", string=email):
    #     messagebox.showwarning(title="Advertencia", message="El correo electronico no es valido")
    #     pass

    # Guardar los datos
    ws.append([nombre, edad, email, telefono, direccion])
    wb.save(nombre_archivo)

    # Limpiar los inputs
    messagebox.showwarning(title="Informacion", message="Datos guardados con exito")
    entry_nombre.delete(first=0, last=tk.END)
    entry_edad.delete(first=0, last=tk.END)
    entry_email.delete(first=0, last=tk.END)
    entry_telefono.delete(first=0, last=tk.END)
    entry_direccion.delete(first=0, last=tk.END)



# Creo la ventana 
root = tk.Tk()
root.title("Formulario de entrada")
root.configure(bg="#4B6587")

# Estilos
label_style = {"bg": "#4B6587", "fg":"white"}
entry_style = {"bg": "#D3D3D3", "fg":"black"}

# labels y inputs del fromulario
# nombre
label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0, column=0,padx=10, pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

# Edad
label_edad = tk.Label(root, text="Edad", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=5)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)

# Email
label_email = tk.Label(root, text="Email", **label_style)
label_email.grid(row=2, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)

# Telefono
label_telefono = tk.Label(root, text="Telefono", **label_style)
label_telefono.grid(row=3, column=0, padx=10, pady=5)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

# Direccion
label_direccion = tk.Label(root, text="Direccion", **label_style)
label_direccion.grid(row=4, column=0, padx=10, pady=5)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=10, pady=5)

# Boton
boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos, bg= "#6D8299", fg="white",width=30)
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=10)



# Para que la ventana no se cierre al ejecutar
root.mainloop()